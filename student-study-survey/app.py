import os
import json
import csv
import uuid
import sqlite3
import io
import requests
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Try importing psycopg2 for cloud PostgreSQL support
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "student_survey_secret_key_2026_xYz987")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "survey.db")
RESPONSES_JSON = os.path.join(DATA_DIR, "responses.json")
RESPONSES_CSV = os.path.join(DATA_DIR, "responses.csv")

ADMIN_PASSCODE = os.environ.get("ADMIN_PASSCODE", "Kushpa45")
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
GOOGLE_SHEET_WEBHOOK_URL = os.environ.get("GOOGLE_SHEET_WEBHOOK_URL", "").strip()

# Normalize postgres:// to postgresql:// if needed
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

os.makedirs(DATA_DIR, exist_ok=True)

CSV_HEADERS = [
    "id",
    "timestamp",
    "student_name",
    "student_email",
    "q1_class",
    "q1_class_other",
    "q2_difficult_subjects",
    "q2_subject_other",
    "q2_why_difficult",
    "q3_biggest_study_problem",
    "q4_when_dont_understand",
    "q4_other_action",
    "q5_how_know_weak_topics",
    "q6_understand_low_marks",
    "q6_other_explanation",
    "q7_teacher_specific_feedback",
    "q8_differentiated_homework",
    "q9_one_month_before_exam",
    "q10_teacher_improvement_wishlist",
    "q11_ai_diagnostic_app_interest",
    "q11_why_interest"
]

def get_pg_connection():
    """Establish connection to cloud PostgreSQL if configured."""
    if DATABASE_URL and HAS_PSYCOPG2:
        try:
            conn = psycopg2.connect(DATABASE_URL, sslmode="prefer")
            return conn
        except Exception as e:
            print(f"PostgreSQL connection error: {e}")
    return None

def init_storage():
    """Initialize Cloud PostgreSQL, SQLite, JSON, and CSV storage."""
    # 1. Cloud PostgreSQL Initialization
    if DATABASE_URL and HAS_PSYCOPG2:
        try:
            conn = get_pg_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS responses (
                        id VARCHAR(64) PRIMARY KEY,
                        timestamp VARCHAR(64),
                        student_name VARCHAR(255),
                        student_email VARCHAR(255),
                        q1_class TEXT,
                        q1_class_other TEXT,
                        q2_difficult_subjects TEXT,
                        q2_subject_other TEXT,
                        q2_why_difficult TEXT,
                        q3_biggest_study_problem TEXT,
                        q4_when_dont_understand TEXT,
                        q4_other_action TEXT,
                        q5_how_know_weak_topics TEXT,
                        q6_understand_low_marks TEXT,
                        q6_other_explanation TEXT,
                        q7_teacher_specific_feedback TEXT,
                        q8_differentiated_homework TEXT,
                        q9_one_month_before_exam TEXT,
                        q10_teacher_improvement_wishlist TEXT,
                        q11_ai_diagnostic_app_interest TEXT,
                        q11_why_interest TEXT
                    );
                ''')
                conn.commit()
                conn.close()
                print("Connected to permanent Cloud PostgreSQL Database!")
                return
        except Exception as e:
            print(f"Cloud PostgreSQL init note: {e}")

    # 2. Local SQLite DB initialization
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS responses (
                id TEXT PRIMARY KEY,
                timestamp TEXT,
                student_name TEXT,
                student_email TEXT,
                q1_class TEXT,
                q1_class_other TEXT,
                q2_difficult_subjects TEXT,
                q2_subject_other TEXT,
                q2_why_difficult TEXT,
                q3_biggest_study_problem TEXT,
                q4_when_dont_understand TEXT,
                q4_other_action TEXT,
                q5_how_know_weak_topics TEXT,
                q6_understand_low_marks TEXT,
                q6_other_explanation TEXT,
                q7_teacher_specific_feedback TEXT,
                q8_differentiated_homework TEXT,
                q9_one_month_before_exam TEXT,
                q10_teacher_improvement_wishlist TEXT,
                q11_ai_diagnostic_app_interest TEXT,
                q11_why_interest TEXT
            )
        ''')
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"SQLite Init Note: {e}")

    # 3. JSON initialization
    if not os.path.exists(RESPONSES_JSON):
        with open(RESPONSES_JSON, "w", encoding="utf-8") as f:
            json.dump([], f, indent=2)

    # 4. CSV initialization
    if not os.path.exists(RESPONSES_CSV):
        with open(RESPONSES_CSV, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(CSV_HEADERS)

init_storage()

def read_responses():
    """Read all responses reliably from Cloud PostgreSQL (or local SQLite/JSON)."""
    # 1. Try Cloud PostgreSQL first
    if DATABASE_URL and HAS_PSYCOPG2:
        try:
            conn = get_pg_connection()
            if conn:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                cursor.execute("SELECT * FROM responses ORDER BY timestamp ASC")
                rows = cursor.fetchall()
                conn.close()

                results = []
                for row in rows:
                    item = dict(row)
                    # Parse JSON array fields if string
                    try:
                        if item.get("q2_difficult_subjects") and item["q2_difficult_subjects"].startswith("["):
                            item["q2_difficult_subjects"] = json.loads(item["q2_difficult_subjects"])
                        elif isinstance(item.get("q2_difficult_subjects"), str):
                            item["q2_difficult_subjects"] = [s.strip() for s in item["q2_difficult_subjects"].split(",") if s.strip()]
                    except Exception:
                        pass

                    try:
                        if item.get("q4_when_dont_understand") and item["q4_when_dont_understand"].startswith("["):
                            item["q4_when_dont_understand"] = json.loads(item["q4_when_dont_understand"])
                        elif isinstance(item.get("q4_when_dont_understand"), str):
                            item["q4_when_dont_understand"] = [s.strip() for s in item["q4_when_dont_understand"].split(",") if s.strip()]
                    except Exception:
                        pass

                    results.append(item)

                return results
        except Exception as e:
            print(f"PostgreSQL read error: {e}")

    # 2. Try SQLite
    try:
        if os.path.exists(DB_PATH):
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM responses ORDER BY rowid ASC")
            rows = cursor.fetchall()
            conn.close()

            results = []
            for row in rows:
                item = dict(row)
                try:
                    if item.get("q2_difficult_subjects") and item["q2_difficult_subjects"].startswith("["):
                        item["q2_difficult_subjects"] = json.loads(item["q2_difficult_subjects"])
                    elif isinstance(item.get("q2_difficult_subjects"), str):
                        item["q2_difficult_subjects"] = [s.strip() for s in item["q2_difficult_subjects"].split(",") if s.strip()]
                except Exception:
                    pass

                try:
                    if item.get("q4_when_dont_understand") and item["q4_when_dont_understand"].startswith("["):
                        item["q4_when_dont_understand"] = json.loads(item["q4_when_dont_understand"])
                    elif isinstance(item.get("q4_when_dont_understand"), str):
                        item["q4_when_dont_understand"] = [s.strip() for s in item["q4_when_dont_understand"].split(",") if s.strip()]
                except Exception:
                    pass

                results.append(item)

            if results:
                return results
    except Exception as e:
        print(f"SQLite read fallback: {e}")

    # 3. Fallback to JSON
    try:
        if os.path.exists(RESPONSES_JSON):
            with open(RESPONSES_JSON, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        print(f"Error reading JSON responses: {e}")
    return []

def save_response(data):
    """Permanently store in Cloud PostgreSQL, SQLite, JSON, and CSV."""
    q2_str = json.dumps(data.get("q2_difficult_subjects", [])) if isinstance(data.get("q2_difficult_subjects"), list) else str(data.get("q2_difficult_subjects", ""))
    q4_str = json.dumps(data.get("q4_when_dont_understand", [])) if isinstance(data.get("q4_when_dont_understand"), list) else str(data.get("q4_when_dont_understand", ""))

    # 1. Save to Cloud PostgreSQL (Permanent Cloud Storage)
    if DATABASE_URL and HAS_PSYCOPG2:
        try:
            conn = get_pg_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO responses (
                        id, timestamp, student_name, student_email,
                        q1_class, q1_class_other,
                        q2_difficult_subjects, q2_subject_other, q2_why_difficult,
                        q3_biggest_study_problem,
                        q4_when_dont_understand, q4_other_action,
                        q5_how_know_weak_topics,
                        q6_understand_low_marks, q6_other_explanation,
                        q7_teacher_specific_feedback,
                        q8_differentiated_homework,
                        q9_one_month_before_exam,
                        q10_teacher_improvement_wishlist,
                        q11_ai_diagnostic_app_interest,
                        q11_why_interest
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        timestamp = EXCLUDED.timestamp,
                        student_name = EXCLUDED.student_name,
                        student_email = EXCLUDED.student_email;
                ''', (
                    data.get("id", ""),
                    data.get("timestamp", ""),
                    data.get("student_name", ""),
                    data.get("student_email", ""),
                    data.get("q1_class", ""),
                    data.get("q1_class_other", ""),
                    q2_str,
                    data.get("q2_subject_other", ""),
                    data.get("q2_why_difficult", ""),
                    data.get("q3_biggest_study_problem", ""),
                    q4_str,
                    data.get("q4_other_action", ""),
                    data.get("q5_how_know_weak_topics", ""),
                    data.get("q6_understand_low_marks", ""),
                    data.get("q6_other_explanation", ""),
                    data.get("q7_teacher_specific_feedback", ""),
                    data.get("q8_differentiated_homework", ""),
                    data.get("q9_one_month_before_exam", ""),
                    data.get("q10_teacher_improvement_wishlist", ""),
                    data.get("q11_ai_diagnostic_app_interest", ""),
                    data.get("q11_why_interest", "")
                ))
                conn.commit()
                conn.close()
                print("Saved permanently to Cloud PostgreSQL!")
        except Exception as e:
            print(f"Cloud PostgreSQL save error: {e}")

    # 2. Save to SQLite (Local DB)
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO responses (
                id, timestamp, student_name, student_email,
                q1_class, q1_class_other,
                q2_difficult_subjects, q2_subject_other, q2_why_difficult,
                q3_biggest_study_problem,
                q4_when_dont_understand, q4_other_action,
                q5_how_know_weak_topics,
                q6_understand_low_marks, q6_other_explanation,
                q7_teacher_specific_feedback,
                q8_differentiated_homework,
                q9_one_month_before_exam,
                q10_teacher_improvement_wishlist,
                q11_ai_diagnostic_app_interest,
                q11_why_interest
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get("id", ""),
            data.get("timestamp", ""),
            data.get("student_name", ""),
            data.get("student_email", ""),
            data.get("q1_class", ""),
            data.get("q1_class_other", ""),
            q2_str,
            data.get("q2_subject_other", ""),
            data.get("q2_why_difficult", ""),
            data.get("q3_biggest_study_problem", ""),
            q4_str,
            data.get("q4_other_action", ""),
            data.get("q5_how_know_weak_topics", ""),
            data.get("q6_understand_low_marks", ""),
            data.get("q6_other_explanation", ""),
            data.get("q7_teacher_specific_feedback", ""),
            data.get("q8_differentiated_homework", ""),
            data.get("q9_one_month_before_exam", ""),
            data.get("q10_teacher_improvement_wishlist", ""),
            data.get("q11_ai_diagnostic_app_interest", ""),
            data.get("q11_why_interest", "")
        ))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"SQLite save error: {e}")

    # 3. Save to JSON
    try:
        responses = read_responses()
        if not any(r.get("id") == data.get("id") for r in responses):
            responses.append(data)
        with open(RESPONSES_JSON, "w", encoding="utf-8") as f:
            json.dump(responses, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"JSON save error: {e}")

    # 4. Append to CSV
    try:
        q2_csv = ", ".join(data.get("q2_difficult_subjects", [])) if isinstance(data.get("q2_difficult_subjects"), list) else str(data.get("q2_difficult_subjects", ""))
        q4_csv = ", ".join(data.get("q4_when_dont_understand", [])) if isinstance(data.get("q4_when_dont_understand"), list) else str(data.get("q4_when_dont_understand", ""))

        row = [
            data.get("id", ""),
            data.get("timestamp", ""),
            data.get("student_name", ""),
            data.get("student_email", ""),
            data.get("q1_class", ""),
            data.get("q1_class_other", ""),
            q2_csv,
            data.get("q2_subject_other", ""),
            data.get("q2_why_difficult", ""),
            data.get("q3_biggest_study_problem", ""),
            q4_csv,
            data.get("q4_other_action", ""),
            data.get("q5_how_know_weak_topics", ""),
            data.get("q6_understand_low_marks", ""),
            data.get("q6_other_explanation", ""),
            data.get("q7_teacher_specific_feedback", ""),
            data.get("q8_differentiated_homework", ""),
            data.get("q9_one_month_before_exam", ""),
            data.get("q10_teacher_improvement_wishlist", ""),
            data.get("q11_ai_diagnostic_app_interest", ""),
            data.get("q11_why_interest", "")
        ]
        with open(RESPONSES_CSV, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(row)
    except Exception as e:
        print(f"CSV save error: {e}")

    # 5. Optional Google Sheets Webhook Sync (Real-time live Google Sheet backup)
    if GOOGLE_SHEET_WEBHOOK_URL:
        try:
            requests.post(GOOGLE_SHEET_WEBHOOK_URL, json=data, timeout=3)
        except Exception as e:
            print(f"Google Sheet webhook sync note: {e}")

def create_styled_portrait_excel(responses):
    """Build a professional, Portrait-formatted Excel sheet with large typography (12-16pt)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Survey Responses"

    # Set Portrait Page Setup for A4 Printing
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Professional Styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1E1B4B")
    sub_font = Font(name="Segoe UI", size=11, italic=True, color="4B5563")
    header_font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=12, bold=False, color="1F2937")
    data_bold = Font(name="Segoe UI", size=12, bold=True, color="111827")

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # 1. Main Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "STUDENT STUDY RESEARCH RESPONSES (PORTRAIT REPORT)"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 2. Subtitle
    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Total Verified Responses: {len(responses)}  |  Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.font = sub_font
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # 3. Column Headers (Clear, descriptive, big font)
    headers = [
        "ID & Time",
        "Student Name",
        "Student Email",
        "Class / Level",
        "Difficult Subject(s) & Reason",
        "Biggest Study Problem",
        "When Stuck on Doubt",
        "How Know Weak Topics",
        "Teacher Feedback & HW",
        "Diagnostic App Interest & Why"
    ]

    ws.append([])
    ws.row_dimensions[3].height = 10

    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 35

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 4. Populate Data Rows with 12pt Large Font & custom height
    current_row = 5
    for idx, r in enumerate(responses):
        ws.row_dimensions[current_row].height = 65

        c_disp = r.get("q1_class", "")
        if str(c_disp).lower() == "above" and r.get("q1_class_other"):
            c_disp = f"Above 12th\n({r.get('q1_class_other')})"

        subjs = r.get("q2_difficult_subjects", [])
        if isinstance(subjs, list):
            subjs_str = ", ".join(subjs)
        else:
            subjs_str = str(subjs or "")
        if r.get("q2_subject_other"):
            subjs_str += f" (Other: {r.get('q2_subject_other')})"
        why_str = f"\nWhy: {r.get('q2_why_difficult', '-')}" if r.get("q2_why_difficult") else ""
        q2_full = f"{subjs_str}{why_str}"

        doubts = r.get("q4_when_dont_understand", [])
        if isinstance(doubts, list):
            doubts_str = ", ".join(doubts)
        else:
            doubts_str = str(doubts or "")
        if r.get("q4_other_action"):
            doubts_str += f" (Other: {r.get('q4_other_action')})"

        t_feedback = f"Feedback: {r.get('q7_teacher_specific_feedback', '-')}\nHW: {r.get('q8_differentiated_homework', '-')}"
        app_interest = f"{r.get('q11_ai_diagnostic_app_interest', '-')}\nWhy: {r.get('q11_why_interest', '-')}"

        row_data = [
            f"#{idx+1} ({r.get('id', '')})\n{r.get('timestamp', '')}",
            r.get("student_name", ""),
            r.get("student_email", ""),
            c_disp,
            q2_full,
            r.get("q3_biggest_study_problem", "-"),
            doubts_str,
            r.get("q5_how_know_weak_topics", "-"),
            t_feedback,
            app_interest
        ]

        row_fill = alt_fill if idx % 2 == 0 else white_fill

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_bold if col_idx in (1, 2) else data_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border

        current_row += 1

    col_widths = {
        1: 18, 2: 20, 3: 24, 4: 16, 5: 32,
        6: 30, 7: 28, 8: 28, 9: 30, 10: 30
    }

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ----------------- ROUTES ----------------- #

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/submit", methods=["POST"])
def submit_survey():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data received"}), 400

        name = data.get("student_name", "").strip()
        email = data.get("student_email", "").strip()

        if not name or not email:
            return jsonify({"success": False, "error": "Name and Email are required."}), 400

        record = {
            "id": str(uuid.uuid4())[:8],
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "student_name": name,
            "student_email": email,
            "q1_class": data.get("q1_class", ""),
            "q1_class_other": data.get("q1_class_other", ""),
            "q2_difficult_subjects": data.get("q2_difficult_subjects", []),
            "q2_subject_other": data.get("q2_subject_other", ""),
            "q2_why_difficult": data.get("q2_why_difficult", ""),
            "q3_biggest_study_problem": data.get("q3_biggest_study_problem", ""),
            "q4_when_dont_understand": data.get("q4_when_dont_understand", []),
            "q4_other_action": data.get("q4_other_action", ""),
            "q5_how_know_weak_topics": data.get("q5_how_know_weak_topics", ""),
            "q6_understand_low_marks": data.get("q6_understand_low_marks", ""),
            "q6_other_explanation": data.get("q6_other_explanation", ""),
            "q7_teacher_specific_feedback": data.get("q7_teacher_specific_feedback", ""),
            "q8_differentiated_homework": data.get("q8_differentiated_homework", ""),
            "q9_one_month_before_exam": data.get("q9_one_month_before_exam", ""),
            "q10_teacher_improvement_wishlist": data.get("q10_teacher_improvement_wishlist", ""),
            "q11_ai_diagnostic_app_interest": data.get("q11_ai_diagnostic_app_interest", ""),
            "q11_why_interest": data.get("q11_why_interest", "")
        }

        save_response(record)
        return jsonify({"success": True, "message": "Response submitted successfully!", "id": record["id"]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/admin")
def admin_page():
    return render_template("admin.html")

@app.route("/api/admin/auth", methods=["POST"])
def admin_auth():
    data = request.get_json() or {}
    passcode = data.get("passcode", "")
    if passcode == ADMIN_PASSCODE:
        session["admin_logged_in"] = True
        return jsonify({"success": True, "message": "Authentication successful"})
    return jsonify({"success": False, "error": "Incorrect passcode"}), 401

@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin_logged_in", None)
    return jsonify({"success": True})

@app.route("/api/admin/data", methods=["GET"])
def admin_data():
    passcode_header = request.headers.get("X-Admin-Passcode")
    if not session.get("admin_logged_in") and passcode_header != ADMIN_PASSCODE:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    responses = read_responses()
    
    total_count = len(responses)
    class_counts = {}
    subject_counts = {}
    doubt_channel_counts = {}
    marks_understanding_counts = {}
    app_interest_counts = {"Yes": 0, "Maybe": 0, "No": 0, "Other": 0}

    for r in responses:
        c = r.get("q1_class") or "Unspecified"
        if str(c).lower() == "above" and r.get("q1_class_other"):
            c = f"Above ({r.get('q1_class_other')})"
        class_counts[c] = class_counts.get(c, 0) + 1

        subjs = r.get("q2_difficult_subjects") or []
        if isinstance(subjs, str):
            subjs = [subjs]
        for s in subjs:
            if str(s).lower() == "other" and r.get("q2_subject_other"):
                s = f"Other ({r.get('q2_subject_other')})"
            subject_counts[s] = subject_counts.get(s, 0) + 1

        channels = r.get("q4_when_dont_understand") or []
        if isinstance(channels, str):
            channels = [channels]
        for ch in channels:
            if "something else" in str(ch).lower() and r.get("q4_other_action"):
                ch = f"Other ({r.get('q4_other_action')})"
            doubt_channel_counts[ch] = doubt_channel_counts.get(ch, 0) + 1

        lm = r.get("q6_understand_low_marks") or "Unspecified"
        if "other" in str(lm).lower() and r.get("q6_other_explanation"):
            lm = f"Other ({r.get('q6_other_explanation')})"
        marks_understanding_counts[lm] = marks_understanding_counts.get(lm, 0) + 1

        ai_int = (r.get("q11_ai_diagnostic_app_interest") or "").lower()
        if "yes" in ai_int:
            app_interest_counts["Yes"] += 1
        elif "maybe" in ai_int:
            app_interest_counts["Maybe"] += 1
        elif "no" in ai_int:
            app_interest_counts["No"] += 1
        else:
            app_interest_counts["Other"] += 1

    analytics = {
        "total_responses": total_count,
        "class_distribution": class_counts,
        "subject_distribution": subject_counts,
        "doubt_channels": doubt_channel_counts,
        "marks_understanding": marks_understanding_counts,
        "app_interest": app_interest_counts
    }

    return jsonify({
        "success": True,
        "responses": list(reversed(responses)),
        "analytics": analytics
    })

@app.route("/api/admin/export/excel")
def export_excel():
    """Download styled Portrait Excel report with large fonts (12-16pt)."""
    passcode = request.args.get("passcode")
    if not session.get("admin_logged_in") and passcode != ADMIN_PASSCODE:
        return jsonify({"error": "Unauthorized"}), 401

    responses = read_responses()
    if not responses:
        return jsonify({"error": "No responses found to export"}), 404

    excel_stream = create_styled_portrait_excel(responses)
    filename = f"Student_Survey_Portrait_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route("/api/admin/print/portrait")
def print_portrait():
    """Render a printable Portrait HTML page with large typography for Save-As-PDF."""
    passcode = request.args.get("passcode")
    if not session.get("admin_logged_in") and passcode != ADMIN_PASSCODE:
        return jsonify({"error": "Unauthorized"}), 401

    responses = read_responses()
    return render_template(
        "portrait_report.html",
        responses=responses,
        date_str=datetime.now().strftime("%d %B %Y, %I:%M %p")
    )

@app.route("/api/admin/export/csv")
def export_csv():
    """Download CSV format with UTF-8 BOM encoding."""
    passcode = request.args.get("passcode")
    if not session.get("admin_logged_in") and passcode != ADMIN_PASSCODE:
        return jsonify({"error": "Unauthorized"}), 401

    responses = read_responses()
    if not responses:
        return jsonify({"error": "No records found"}), 404

    # Build fresh CSV from current database responses
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_HEADERS)

    for r in responses:
        q2_csv = ", ".join(r.get("q2_difficult_subjects", [])) if isinstance(r.get("q2_difficult_subjects"), list) else str(r.get("q2_difficult_subjects", ""))
        q4_csv = ", ".join(r.get("q4_when_dont_understand", [])) if isinstance(r.get("q4_when_dont_understand"), list) else str(r.get("q4_when_dont_understand", ""))
        writer.writerow([
            r.get("id", ""),
            r.get("timestamp", ""),
            r.get("student_name", ""),
            r.get("student_email", ""),
            r.get("q1_class", ""),
            r.get("q1_class_other", ""),
            q2_csv,
            r.get("q2_subject_other", ""),
            r.get("q2_why_difficult", ""),
            r.get("q3_biggest_study_problem", ""),
            q4_csv,
            r.get("q4_other_action", ""),
            r.get("q5_how_know_weak_topics", ""),
            r.get("q6_understand_low_marks", ""),
            r.get("q6_other_explanation", ""),
            r.get("q7_teacher_specific_feedback", ""),
            r.get("q8_differentiated_homework", ""),
            r.get("q9_one_month_before_exam", ""),
            r.get("q10_teacher_improvement_wishlist", ""),
            r.get("q11_ai_diagnostic_app_interest", ""),
            r.get("q11_why_interest", "")
        ])

    mem = io.BytesIO()
    mem.write(output.getvalue().encode('utf-8-sig'))
    mem.seek(0)

    return send_file(
        mem,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"student_survey_responses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    )

@app.route("/api/admin/export/json")
def export_json():
    passcode = request.args.get("passcode")
    if not session.get("admin_logged_in") and passcode != ADMIN_PASSCODE:
        return jsonify({"error": "Unauthorized"}), 401

    responses = read_responses()
    mem = io.BytesIO()
    mem.write(json.dumps(responses, indent=2, ensure_ascii=False).encode('utf-8'))
    mem.seek(0)

    return send_file(
        mem,
        mimetype="application/json",
        as_attachment=True,
        download_name=f"student_survey_responses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("\n" + "="*60)
    print("Student Study Research Survey Server Started!")
    print(f"Student Survey URL: http://localhost:{port}")
    print(f"Researcher Admin Portal: http://localhost:{port}/admin")
    print("="*60 + "\n")
    app.run(host="0.0.0.0", port=port, debug=True)
